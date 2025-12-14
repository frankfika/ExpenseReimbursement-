"""报表生成模块 - 生成 Excel 统计报表"""
from pathlib import Path
from typing import Dict, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from invoice_analyzer import InvoiceInfo
from config import INVOICE_CATEGORIES


class ReportGenerator:
    """报表生成器"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)

    def _sanitize_sheet_name(self, name: str) -> str:
        """清理工作表名称，移除Excel不允许的字符"""
        # Excel 不允许: : / \ ? * [ ]
        invalid_chars = [':', '/', '\\', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        # 工作表名称最长31个字符
        return name[:31]

    def generate(self, categorized: Dict[str, List[InvoiceInfo]]) -> str:
        """
        生成报销统计报表

        Args:
            categorized: 分类后的发票信息 {类别名: [发票信息列表]}

        Returns:
            报表文件路径
        """
        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 1. 先创建各分类明细表（记录每个表的行数，用于汇总表公式）
        sheet_info = {}  # {分类名: (工作表名, 发票数量)}
        for category_name, infos in categorized.items():
            if infos:
                # 只统计发票数量
                invoice_count = len([i for i in infos if i.is_invoice])
                if invoice_count > 0:
                    sheet_name = self._create_detail_sheet(wb, category_name, infos)
                    sheet_info[category_name] = (sheet_name, invoice_count)

        # 2. 创建汇总表（使用公式引用明细表）
        self._create_summary_sheet_with_formulas(wb, sheet_info)

        # 保存文件（固定文件名，每次覆盖）
        report_path = self.output_dir / "报销统计.xlsx"
        wb.save(str(report_path))

        return str(report_path)

    def _create_summary_sheet(self, wb: Workbook, categorized: Dict[str, List[InvoiceInfo]]):
        """创建汇总表"""
        ws = wb.create_sheet("汇总")

        # 定义样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        # 标题
        ws.merge_cells('A1:D1')
        ws['A1'] = f"报销汇总表 - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_align

        # 表头
        headers = ['类别', '发票数量', '金额（元）', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # 数据行
        total_count = 0
        total_amount = 0.0
        row = 4

        # 按预定义顺序排列类别
        category_order = ['打车票', '火车飞机票', '住宿费', '餐费', '其他']
        for category_name in category_order:
            if category_name in categorized:
                infos = categorized[category_name]
                # 只统计发票金额
                invoice_infos = [i for i in infos if i.is_invoice]
                count = len(invoice_infos)
                amount = sum(i.amount for i in invoice_infos)

                ws.cell(row=row, column=1, value=category_name).border = border
                ws.cell(row=row, column=2, value=count).border = border
                ws.cell(row=row, column=2).alignment = center_align
                cell = ws.cell(row=row, column=3, value=amount)
                cell.border = border
                cell.alignment = right_align
                cell.number_format = '#,##0.00'
                ws.cell(row=row, column=4, value="").border = border

                total_count += count
                total_amount += amount
                row += 1

        # 合计行
        ws.cell(row=row, column=1, value="合计").font = header_font
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=total_count).font = header_font
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).alignment = center_align
        cell = ws.cell(row=row, column=3, value=total_amount)
        cell.font = header_font
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        ws.cell(row=row, column=4, value="").border = border

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

    def _create_detail_sheet(self, wb: Workbook, category_name: str, infos: List[InvoiceInfo]) -> str:
        """创建分类明细表，返回工作表名称（只包含发票，不包含凭证）"""
        sheet_name = self._sanitize_sheet_name(category_name)
        ws = wb.create_sheet(sheet_name)

        # 只保留发票，过滤掉凭证
        invoice_infos = [i for i in infos if i.is_invoice]

        # 如果没有发票，返回空表
        if not invoice_infos:
            ws.cell(row=1, column=1, value="暂无发票")
            return sheet_name

        # 定义样式
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 表头（移除"类型"列，因为都是发票）
        headers = ['序号', '日期', '商家/平台', '金额（元）', '发票号码', '描述', '文件路径']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # 数据行
        row = 2
        total_amount = 0.0

        # 按日期排序（只处理发票）
        sorted_infos = sorted(invoice_infos, key=lambda x: x.date or "")

        for idx, info in enumerate(sorted_infos, 1):
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center_align

            ws.cell(row=row, column=2, value=info.date).border = border
            ws.cell(row=row, column=2).alignment = center_align

            ws.cell(row=row, column=3, value=info.subtype or info.merchant).border = border

            cell = ws.cell(row=row, column=4, value=info.amount)
            cell.border = border
            cell.alignment = right_align
            cell.number_format = '#,##0.00'

            ws.cell(row=row, column=5, value=info.invoice_number).border = border

            ws.cell(row=row, column=6, value=info.description).border = border
            ws.cell(row=row, column=6).alignment = wrap_align

            # 相对路径
            try:
                rel_path = Path(info.file_path).relative_to(self.output_dir)
            except ValueError:
                rel_path = info.file_path
            ws.cell(row=row, column=7, value=str(rel_path)).border = border

            total_amount += info.amount
            row += 1

        # 合计行 - 使用 SUM 公式
        ws.cell(row=row, column=1, value="").border = border
        ws.cell(row=row, column=2, value="").border = border
        ws.cell(row=row, column=3, value="合计").font = header_font
        ws.cell(row=row, column=3).border = border
        # 直接用 SUM 公式，因为都是发票
        sum_formula = f'=SUM(D2:D{row-1})'
        cell = ws.cell(row=row, column=4, value=sum_formula)
        cell.font = header_font
        cell.border = border
        cell.alignment = right_align
        cell.number_format = '#,##0.00'
        for col in range(5, 8):
            ws.cell(row=row, column=col, value="").border = border

        # 调整列宽
        col_widths = [6, 12, 20, 12, 20, 25, 40]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 隐藏文件路径列（第7列）
        ws.column_dimensions['G'].hidden = True

        return sheet_name

    def _create_summary_sheet_with_formulas(self, wb: Workbook, sheet_info: Dict[str, tuple]):
        """创建汇总表，使用公式引用明细表"""
        ws = wb.create_sheet("汇总", 0)  # 插入到第一个位置

        # 定义样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        # 标题
        ws.merge_cells('A1:D1')
        ws['A1'] = f"报销汇总表 - {datetime.now().strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_align

        # 表头
        headers = ['类别', '发票数量', '金额（元）', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # 数据行 - 使用公式引用明细表
        row = 4
        count_cells = []  # 记录数量单元格，用于合计
        amount_cells = []  # 记录金额单元格，用于合计

        # 按预定义顺序排列类别
        category_order = ['打车票', '火车飞机票', '住宿费', '餐费', '其他']
        for category_name in category_order:
            if category_name in sheet_info:
                sheet_name, data_rows = sheet_info[category_name]
                last_data_row = data_rows + 1  # 数据从第2行开始

                ws.cell(row=row, column=1, value=category_name).border = border

                # 发票数量公式：统计有数据的行数（用 COUNTA 统计 A 列，减去表头）
                count_formula = f"=COUNTA('{sheet_name}'!A2:A{last_data_row})"
                count_cell = ws.cell(row=row, column=2, value=count_formula)
                count_cell.border = border
                count_cell.alignment = center_align
                count_cells.append(f"B{row}")

                # 金额公式：使用明细表的合计单元格（在最后一行的D列）
                amount_formula = f"='{sheet_name}'!D{last_data_row + 1}"
                amount_cell = ws.cell(row=row, column=3, value=amount_formula)
                amount_cell.border = border
                amount_cell.alignment = right_align
                amount_cell.number_format = '#,##0.00'
                amount_cells.append(f"C{row}")

                ws.cell(row=row, column=4, value="").border = border
                row += 1

        # 合计行 - 使用 SUM 公式
        ws.cell(row=row, column=1, value="合计").font = header_font
        ws.cell(row=row, column=1).border = border

        # 数量合计
        if count_cells:
            count_sum = f"=SUM({','.join(count_cells)})"
        else:
            count_sum = 0
        count_total_cell = ws.cell(row=row, column=2, value=count_sum)
        count_total_cell.font = header_font
        count_total_cell.border = border
        count_total_cell.alignment = center_align

        # 金额合计
        if amount_cells:
            amount_sum = f"=SUM({','.join(amount_cells)})"
        else:
            amount_sum = 0
        amount_total_cell = ws.cell(row=row, column=3, value=amount_sum)
        amount_total_cell.font = header_font
        amount_total_cell.border = border
        amount_total_cell.alignment = right_align
        amount_total_cell.number_format = '#,##0.00'

        ws.cell(row=row, column=4, value="").border = border

        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20


def generate_report(output_dir: str, categorized: Dict[str, List[InvoiceInfo]]) -> str:
    """便捷函数：生成报表"""
    generator = ReportGenerator(output_dir)
    return generator.generate(categorized)
